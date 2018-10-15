import threading

import pymysql
from openpyxl import load_workbook

from decorator.timekill import clock


class myThread (threading.Thread):
    def __init__(self, threadID, name, sqlcontent , colnumber):
        threading.Thread.__init__(self)
        self.threadID = threadID
        self.name = name
        self.sqlcontent = sqlcontent
        self.colnumber = colnumber


    def run(self):
        print_time(self.sqlcontent, self.colnumber)

def print_time( sqlcontent, colnumber):

    # 使用 execute()  方法执行 SQL 查询
    cursor.execute("SELECT * FROM `"+sqlcontent+"`")
    results = cursor.fetchall()

    # 给一个默认的游标
    n = 2
    for data in results:
        ws.cell(n, colnumber, data[0])
        n += 1

# 加载对应的excel文件
wb = load_workbook('C:\\Users\\Administrator\\Desktop\\higherT.xlsx')
ws = wb.active

# 打开数据库连接
db = pymysql.connect("localhost", "root", "root", "formativedata")
# 使用 cursor() 方法创建一个游标对象 cursor
cursor = db.cursor()

@clock
def testtime():
    # 初始化
    thread1 = myThread(1, "Thread-1", "出院科室名称",8)
    thread2 = myThread(2, "Thread-2", "入院日期",9)
    thread3 = myThread(3, "Thread-3", "出院日期", 10)
    thread4 = myThread(4, "Thread-4", "主要诊断", 11)
    thread5 = myThread(5, "Thread-5", "主要诊断名称", 12)
    thread6 = myThread(6, "Thread-6", "病人婚姻", 13)

    # 执行线程
    thread1.start()
    thread1.join()
    thread2.start()
    thread2.join()
    thread3.start()
    thread3.join()
    thread4.start()
    thread4.join()
    thread5.start()
    thread5.join()
    thread6.start()
    thread6.join()


testtime()

# 最后，将以上操作保存到指定的Excel文件中
wb.save('C:\\Users\\Administrator\\Desktop\\higherT.xlsx')

print("退出主线程")