
#打开文件：
import pymysql
from openpyxl import load_workbook

# 加载对应的excel文件
from decorator.timekill import clock

wb=load_workbook('C:\\Users\\Administrator\\Desktop\\hasNullData.xlsx')
ws=wb.active

# 打开数据库连接
db = pymysql.connect("localhost", "root", "root", "formativedata")
# 使用 cursor() 方法创建一个游标对象 cursor
cursor = db.cursor()

@clock
def testtime():

    cursor.execute("SELECT * FROM `就诊id`")
    results = cursor.fetchall()

    cursor.execute("SELECT * FROM `采样时间`")
    results1 = cursor.fetchall()

    cursor.execute("SELECT * FROM `标本种类`")
    results2 = cursor.fetchall()

    cursor.execute("SELECT * FROM `采集部位`")
    results3 = cursor.fetchall()

    cursor.execute("SELECT * FROM `审核时间`")
    results4 = cursor.fetchall()

    cursor.execute("SELECT * FROM `项目id`")
    results5 = cursor.fetchall()

    cursor.execute("SELECT * FROM `项目名称`")
    results6 = cursor.fetchall()

    cursor.execute("SELECT * FROM `定性结果`")
    results7 = cursor.fetchall()

    cursor.execute("SELECT * FROM `结果`")
    results8 = cursor.fetchall()

    cursor.execute("SELECT * FROM `参考范围`")
    results9 = cursor.fetchall()

    cursor.execute("SELECT * FROM `单位`")
    results10 = cursor.fetchall()

    cursor.execute("SELECT * FROM `检测方法`")
    results11 = cursor.fetchall()

    cursor.execute("SELECT * FROM `检测仪器`")
    results12 = cursor.fetchall()

    n=2
    for data in results:
        ws.cell(n, 44, data[0])
        n+=1

    n=2
    for data in results1:
        ws.cell(n, 45, data[0])
        n+=1

    n = 2
    for data in results2:
        ws.cell(n, 46, data[0])
        n += 1

    n = 2
    for data in results3:
        ws.cell(n, 47, data[0])
        n += 1

    n = 2
    for data in results4:
        ws.cell(n, 48, data[0])
        n += 1

    n = 2
    for data in results5:
        ws.cell(n, 49, data[0])
        n += 1

    n = 2
    for data in results6:
        ws.cell(n, 50, data[0])
        n += 1

    n = 2
    for data in results7:
        ws.cell(n, 51, data[0])
        n += 1

    n = 2
    for data in results8:
        ws.cell(n, 52, data[0])
        n += 1

    n = 2
    for data in results9:
        ws.cell(n, 53, data[0])
        n += 1

    n = 2
    for data in results10:
        ws.cell(n, 54, data[0])
        n += 1

    n = 2
    for data in results11:
        ws.cell(n, 55, data[0])
        n += 1

    n = 2
    for data in results12:
        ws.cell(n, 56, data[0])
        n += 1

testtime()

# 关闭数据库连接
db.close()

#最后，将以上操作保存到指定的Excel文件中
wb.save('C:\\Users\\Administrator\\Desktop\\hasNullData.xlsx')