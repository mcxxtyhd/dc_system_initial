import threading

import pymysql
from openpyxl import load_workbook

class myThread (threading.Thread):
    def __init__(self, threadID, name, sqlcontent , colnumber):
        threading.Thread.__init__(self)
        self.threadID = threadID
        self.name = name
        self.sqlcontent = sqlcontent
        self.colnumber = colnumber


    def run(self):
        print_time(self.sqlcontent, self.colnumber)

def print_time( sqlcontent, colnumber):

    # 使用 execute()  方法执行 SQL 查询
    cursor.execute("SELECT * FROM `"+sqlcontent+"`")
    results = cursor.fetchall()

    # 给一个默认的游标
    n = 2
    for data in results:
        ws.cell(n, colnumber, data[0])
        n += 1

# 加载对应的excel文件
wb = load_workbook('C:\\Users\\Administrator\\Desktop\\higher.xlsx')
ws = wb.active

# 打开数据库连接
db = pymysql.connect("localhost", "root", "root", "formativedata")
# 使用 cursor() 方法创建一个游标对象 cursor
cursor = db.cursor()

# 初始化
thread1 = myThread(1, "Thread-1", "出院科室名称",8)
thread2 = myThread(2, "Thread-2", "入院日期",9)

# 执行线程
thread1.start()
thread1.join()
thread2.start()
thread2.join()

# 最后，将以上操作保存到指定的Excel文件中
wb.save('C:\\Users\\Administrator\\Desktop\\higher.xlsx')

print("退出主线程")